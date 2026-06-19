import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__, static_folder=".", static_url_path="")

ORDERS_FILE = "commandes.xlsx"
ADMIN_EMAIL = "info@yourmealbox.com"


def init_workbook():
    if not os.path.exists(ORDERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commandes"
        ws.append(["Date", "Prénom", "Email", "Formule", "Jour de livraison", "Message"])
        wb.save(ORDERS_FILE)


def save_order(data):
    try:
        init_workbook()
        wb = load_workbook(ORDERS_FILE)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            data.get("prenom", ""),
            data.get("email", ""),
            data.get("formule", ""),
            data.get("jourLivraison", ""),
            data.get("message", ""),
        ])
        wb.save(ORDERS_FILE)
    except Exception as e:
        print(f"Excel save error (non-fatal): {e}")


def send_email(data):
    email_user = os.getenv("EMAIL_USER")
    email_pass = os.getenv("EMAIL_PASS")
    if not email_user or not email_pass:
        print("EMAIL_USER or EMAIL_PASS not set — skipping email")
        return False

    body = f"""Nouvelle commande YourMealBox!

Prénom : {data.get('prenom', '')}
Email : {data.get('email', '')}
Formule : {data.get('formule', '')}
Jour de livraison : {data.get('jourLivraison', '')}
Message : {data.get('message', '')}

Date : {datetime.now().strftime("%Y-%m-%d %H:%M")}
"""

    msg = MIMEMultipart()
    msg["From"] = email_user
    msg["To"] = ADMIN_EMAIL
    msg["Subject"] = f"Nouvelle commande – {data.get('prenom', 'Client')}"
    msg.attach(MIMEText(body, "plain", "utf-8"))

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(email_user, email_pass)
            server.sendmail(email_user, ADMIN_EMAIL, msg.as_string())
        print(f"Email sent to {ADMIN_EMAIL}")
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False


@app.route("/")
def index():
    return send_from_directory(".", "index.html")


@app.route("/api/commande", methods=["POST"])
def commande():
    data = request.get_json(force=True, silent=True) or {}
    print(f"Received order: prenom={data.get('prenom')}, email={data.get('email')}, formule={data.get('formule')}")

    if not data.get("prenom") or not data.get("email"):
        return jsonify({"error": "Prénom et email requis"}), 400

    save_order(data)
    email_sent = send_email(data)
    print(f"Order processed. Email sent: {email_sent}")

    return jsonify({"success": True, "message": "Commande reçue!"}), 200


if __name__ == "__main__":
    port = int(os.getenv("PORT", 3000))
    app.run(host="0.0.0.0", port=port, debug=False)
